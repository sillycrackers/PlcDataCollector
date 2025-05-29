import os
import shutil
import sys
import traceback
import winreg
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

import src.plc_connection as pc
from src.ticketing_system import TicketPurpose, Ticket, transmit


def copy_paste_file(file_path, dest_path):

    shutil.copy(file_path, dest_path)

def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS

    except Exception:
        base_path = os.path.abspath("..")

    return os.path.join(base_path, relative_path)

#Save the last .pdc file open to Windows register, so we can load it next time the software opens
def set_reg(file_path):

    try:
        software = winreg.OpenKeyEx(winreg.HKEY_CURRENT_USER, r"SOFTWARE\\")
        key = winreg.CreateKeyEx(software, "Plc Data Collector")
        winreg.SetValueEx(key, "last_file_path", 0, winreg.REG_SZ, file_path)
        key.Close()
    except Exception:
        print("ERROR: Couldn't change Windows Registry")

#Load the last file open from Windows registry
def get_reg(reg_path):

    try:
        pdc = winreg.OpenKey(winreg.HKEY_CURRENT_USER, reg_path)
        file_path_regval = winreg.QueryValueEx(pdc, "last_file_path")

        if file_path_regval:
            pdc.Close()
            return file_path_regval[0]

    except (FileNotFoundError, WindowsError):
        print(f"Couldn't find '{reg_path}\\last_file_path' in registry")
        return False

# Function to save data to Excel
def save_tag_data_to_excel(plc, data_row, main_frame, write_type):

    # If folder doesn't exist, then create it
    if not os.path.exists(plc.excel_file_location):
        os.makedirs(plc.excel_file_location)
    if data_row:
        try:
            save_data_to_excel(headers=plc.tags, data_row=data_row, file_path=plc.file_path, sheet_name="Plc Data", write_type=write_type)
            transmit(main_frame,Ticket(purpose=TicketPurpose.OUTPUT_MESSAGE, value=f"Data collected from {plc.name}: {data_row}"))

        except:
            traceback.print_exc()

def adjust_column_width(ws : Worksheet):

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].auto_size = True

def save_data_to_excel(headers, data_row, file_path, sheet_name, write_type):
    # If file already exists then open it
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        ws = wb.active

    # If Excel file doesn't exist then create it
    else:
        # Create excel workbook, take active sheet and name it PLC Data
        # Create initial column headers named: "Timestamp" followed by the tag names
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

    overwrite_row(ws=ws,row=1,data=["Timestamp"] + headers)

    if write_type==pc.WriteType.APPEND:
        # Add plc tag data as the next row in the sheet
        ws.append(data_row)

    elif write_type==pc.WriteType.OVERWRITE:
        # Overwrite second row after headers with new values
        overwrite_row(ws=ws, row=2, data=data_row)

    # Try to autosize the width of the columns
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].auto_size = True

    # Save and close Excel file after logging the tag data
    wb.save(file_path)

def archive(data_list : [[]], archive_file_location : str, archive_file_name : str, headers : []):

    file_path = f"{archive_file_location}\\{archive_file_name}.xlsx"

    try:
        os.makedirs(archive_file_location)
    except FileExistsError:
        print(f"Archive folder already exists: {archive_file_location}")

    try:
        # Create excel workbook, take active sheet and name it PLC Data
        # Create initial column headers named: "Timestamp" followed by the tag names
        wb = Workbook()
        ws = wb.active
        ws.title = "archive"
        first_header = "Timestamp"

        ws.append([first_header] + headers)  # Create Header row

        for data_row in data_list:
            ws.append(data_row)


        # Try to autosize the width of the columns
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].auto_size = True

        # Save and close Excel file after logging the tag data
        wb.save(file_path)
        return True

    except Exception:
        print(f"Error trying to save archive file {archive_file_name}")
        return False




def overwrite_row(ws, row, data):

    for col_num, d in enumerate(data, start=1):
        ws.cell(row=row, column=col_num).value = d


