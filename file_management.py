import os
import shutil
import sys
import traceback
import winreg
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.worksheet.worksheet import Worksheet

from ticketing_system import *


def copy_paste_file(file_path, dest_path):

    shutil.copy(file_path, dest_path)

def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS

    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

def set_reg(file_path):

    try:
        software = winreg.OpenKeyEx(winreg.HKEY_CURRENT_USER, r"SOFTWARE\\")
        key = winreg.CreateKeyEx(software, "Plc Data Collector")
        winreg.SetValueEx(key, "last_file_path", 0, winreg.REG_SZ, file_path)
        key.Close()
    except Exception:
        print("ERROR: Couldn't change Windows Registry")

def get_reg(reg_path):

    try:
        pdc = winreg.OpenKey(winreg.HKEY_CURRENT_USER, reg_path)
        file_path_regval = winreg.QueryValueEx(pdc, "last_file_path")

        if file_path_regval:
            pdc.Close()
            return file_path_regval[0]

    except (FileNotFoundError, WindowsError):
        print(f"Couldn't find '{reg_path}\\last_file_path' in registry")
        return False

# Function to save data to Excel
def save_to_excel(plc, row, ticketer : TicketingSystem):

    # If folder doesn't exist, then create it
    if not os.path.exists(plc.excel_file_location):
        os.makedirs(plc.excel_file_location)

    if row:
        try:
            # If file already exists then open it
            if os.path.exists(plc.file_path):
                wb = load_workbook(plc.file_path)
                ws = wb.active
            # If Excel file doesn't exist then create it
            else:
                # Create excel workbook, take active sheet and name it PLC Data
                # Create initial column headers named: "Timestamp" followed by the tag names
                wb = Workbook()
                ws = wb.active

                ws.title = "PLC Data"
                ws.append(["Timestamp"] + plc.tags)  # Header row

            # Add plc tag data as the next row in the sheet
            ws.append(row)

            for col in ws.columns:
                print("Test")
                ws.column_dimensions[col[0].column_letter].auto_size = True

            # Save and close excel file after logging the tag data
            wb.save(plc.file_path)

            ticketer.transmit(Ticket(purpose=TicketPurpose.OUTPUT_MESSAGE, value=f"Data collected from {plc.name}: {row}"))

        except:
            traceback.print_exc()

def adjust_column_width(ws : Worksheet):

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].auto_size = True