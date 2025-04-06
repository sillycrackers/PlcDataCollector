import os
import sys
import traceback
import winreg
import ttkbootstrap as ttk
import openpyxl
from openpyxl import Workbook, load_workbook
from enum import Enum, auto
import shutil

import gui.prompt


def copy_paste_file(file_path, dest_path):

    shutil.copy(file_path, dest_path)

def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS

    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

def set_reg(file_path):

    try:
        software = winreg.OpenKeyEx(winreg.HKEY_CURRENT_USER,r"SOFTWARE\\")
        key = winreg.CreateKeyEx(software, "Plc Data Collector")
        winreg.SetValueEx(key, "last_file_path",0, winreg.REG_SZ, file_path)
        key.Close()
    except Exception:
        print("ERROR: Couldn't change Windows Registry")

def get_reg(reg_path):

    try:
        pdc = winreg.OpenKey(winreg.HKEY_CURRENT_USER, reg_path)
        file_path_regval = winreg.QueryValueEx(pdc,"last_file_path")

        if file_path_regval:
            pdc.Close()
            return file_path_regval[0]

    except (FileNotFoundError, WindowsError):
        print("Couldn't find 'SOFTWARE\\Plc Data Collector\\last_file_path' in registry")
        return False

def change_theme(theme):
    if theme == 'dark':

        style_object = ttk.Style(theme='darkly')

        style_object.configure('TLabelframe.Label', font=('Calibri', 12,))
        style_object.configure('custom.TButton', font=('Calibri', 12,))
        style_object.configure(style='alarm.Treeview', font=('Calibri', 12,), foreground="red")

    elif theme == 'light':

        style_object = ttk.Style(theme='flatly')

        style_object.configure('TLabelframe.Label', font=('Calibri', 12,))
        style_object.configure('custom.TButton', font=('Calibri', 12,))
        style_object.configure(style='alarm.Treeview', font=('Calibri', 12,), foreground="red")

# Function to save data to Excel
def save_to_excel(plc, row):

    #If folder doesn't exist, then create it
    if not os.path.exists(plc.excel_file_location):
        os.makedirs(plc.excel_file_location)

    if row:
        try:
            #If file already exists then open it
            if os.path.exists(plc.file_path):
                wb = load_workbook(plc.file_path)
                ws = wb.active
            #If Excel file doesn't exist then create it
            else:
                #Create excel workbook, take active sheet and name it PLC Data
                #Create initial column headers named: "Timestamp" followed by the tag names
                wb = Workbook()
                ws = wb.active
                ws.title = "PLC Data"
                ws.append(["Timestamp"] + plc.tags)  # Header row

            #Add plc tag data as the next row in the sheet
            ws.append(row)
            #Save and close excel file after logging the tag data
            wb.save(plc.file_path)
            print(f"Data logged: {row} to {plc.file_path}")
        except:
            traceback.print_exc()

class TicketPurpose(Enum):

    # ("message":str, Alarm active:bool)
    UPDATE_ALARMS = auto()
    # (state:bool,"plc.name:str")
    TOGGLE_INDICATOR = auto()

    POPULATE_INDICATORS = auto()

    ACTIVE_ALARMS_CLEAR = auto()

    SHOW_WAIT_CURSOR = auto()

    SHOW_NORMAL_CURSOR = auto()

    # (AnimatedLabel: object,column : int, row : int)
    SHOW_ANIMATED_LABEL = auto()
    # (AnimatedLabel: object,column : int, row : int)
    HIDE_ANIMATED_LABEL = auto()

    OUTPUT_MESSAGE = auto()

class Ticket:
    def __init__(self, purpose: TicketPurpose, value, main_frame):
        self.purpose = purpose
        self.value = value
        self.main_frame = main_frame

    def transmit(self):

        self.main_frame.q.put(self)
        self.main_frame.event_generate("<<CheckQueue>>")

def disable_event(parent):
   gui.prompt.Prompt(parent)
